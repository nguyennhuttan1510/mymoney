import os.path
from abc import ABC, abstractmethod
from datetime import datetime

from django.http import HttpResponse
from openpyxl.reader.excel import load_workbook
from openpyxl.workbook import Workbook
from openpyxl.worksheet.worksheet import Worksheet
from typing_extensions import TypeVar, Generic, List

from report.schema import ReportOut

T = TypeVar('T')

class ReportTemplateAbstract(Generic[T], ABC):
    def open_file_template(self, path: str):
        pass

    @abstractmethod
    def handle_data(self):
        pass

    @abstractmethod
    def export(self):
        pass



class TransactionReportTemplate(ReportTemplateAbstract):
    TEMPLATE_PATH = 'report_transaction.xlsx'
    wb: Workbook
    ws: Worksheet
    def __init__(self, data: ReportOut):
        self.data = data
        self.open_file_template(self.TEMPLATE_PATH)
        self.handle_data()

    def export(self):
        response = HttpResponse(
            content_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
        )
        filename = f"transactions_report_{datetime.now().strftime('%Y%m%d_%H%M%S')}.xlsx"
        response["Content-Disposition"] = f"attachment; filename={filename}"
        self.wb.save(response)
        return response


    def open_file_template(self, file_name: str):
        TEMPLATE_PATH = os.path.join('templates', file_name)
        self.wb = load_workbook(TEMPLATE_PATH)
        # self.ws.title = 'Transaction Report'

    def _handle_data_transaction(self):
        self.ws = self.wb.worksheets[1]

        start_row = 5
        for idx, tx in enumerate(self.data.transactions, start=start_row):
            self.ws.cell(row=idx, column=2).value = tx.transaction_date.strftime("%d/%m/%Y %H:%M")
            self.ws.cell(row=idx, column=3).value = tx.amount
            self.ws.cell(row=idx, column=4).value = tx.note
            self.ws.cell(row=idx, column=5).value = tx.category.name

    def _handle_data_overview(self):
        self.ws = self.wb.worksheets[0]

        start_row = 28
        for idx, category in enumerate(self.data.categories, start=start_row):
            self.ws.cell(row=idx, column=2).value = category.name
            self.ws.cell(row=idx, column=5).value = category.total

    def handle_data(self):
        # headers = ['ID', 'Wallet', 'Category', 'Amount', 'Date']
        # self.ws.append(headers)
        self._handle_data_overview()
        self._handle_data_transaction()




